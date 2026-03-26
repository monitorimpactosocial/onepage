from pathlib import Path
from datetime import datetime
import shutil
import requests
from flask import current_app
from openpyxl import load_workbook
from ..models import Question, Submission, Response, OptionItem, Attachment, AppSetting

DATA_SHEET_MAP = {
    "ENVIOS": ["submission_key","period_key","area_code","status","responsible_user","opened_at","saved_at","submitted_at","completeness_pct","reminder_sent_at","comments"],
    "RESPUESTAS": ["id","submission_key","period_key","period_date","area_code","form_code","section","group_key","row_key","question_key","label","value_type","value_num","value_text","value_date","value_bool","unit","observation","created_by","updated_at","source_channel"],
    "ADJUNTOS": ["attachment_id","submission_key","period_key","area_code","original_filename","stored_filename","content_type","uploaded_by","uploaded_at"],
    "NARRATIVAS": ["submission_key","period_key","area_code","responsible_user","narrative_text","comments"],
    "CAT_LISTAS": ["list_code","value_code","value_label","group_name","owner_area_code","user_editable","sort_order","is_active","created_by"],
    "CAT_SETTINGS": ["setting_key","setting_value","updated_at"],
}

def _clear_body(ws, start_row=5):
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)

def _col_letter(n):
    letters = ""
    while n:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters

def _write_rows(ws, headers, rows, start_row=4):
    for idx, head in enumerate(headers, start=1):
        ws.cell(start_row, idx, head)
    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)
    if ws.tables:
        name = list(ws.tables)[0]
        last_col = _col_letter(len(headers))
        last_row = max(start_row + len(rows), start_row + 1)
        ws.tables[name].ref = f"A{start_row}:{last_col}{last_row}"

def _response_rows():
    rows = []
    questions = {q.question_key: q for q in Question.query.all()}
    for resp in Response.query.join(Submission).all():
        q = questions.get(resp.question_key)
        value_type = ""
        if resp.value_num is not None:
            value_type = "number"
        elif resp.value_text:
            value_type = "text"
        elif resp.value_date:
            value_type = "date"
        elif resp.value_bool is not None:
            value_type = "bool"
        rows.append([
            resp.id,
            resp.submission.submission_key,
            resp.submission.period_key,
            datetime.strptime(resp.submission.period_key + "-01", "%Y-%m-%d"),
            resp.submission.area_code,
            q.form_code if q else "",
            q.section if q else "",
            q.group_key if q else "",
            resp.row_key,
            resp.question_key,
            q.label if q else resp.question_key,
            value_type,
            resp.value_num,
            resp.value_text,
            resp.value_date,
            resp.value_bool,
            resp.unit,
            resp.observation,
            resp.created_by,
            resp.updated_at,
            "App web"
        ])
    return rows

def _submission_rows():
    return [[
        s.submission_key, s.period_key, s.area_code, s.status, s.responsible_user, s.opened_at, s.saved_at,
        s.submitted_at, (s.completeness_pct or 0) / 100 if (s.completeness_pct or 0) > 1 else s.completeness_pct,
        s.reminder_sent_at, s.comments
    ] for s in Submission.query.all()]

def _attachment_rows():
    rows = []
    for a in Attachment.query.join(Submission).all():
        rows.append([
            a.id, a.submission.submission_key, a.submission.period_key, a.submission.area_code,
            a.original_filename, a.stored_filename, a.content_type, a.uploaded_by, a.uploaded_at
        ])
    return rows

def _narrative_rows():
    return [[s.submission_key, s.period_key, s.area_code, s.responsible_user, s.narrative_text, s.comments] for s in Submission.query.all()]

def _list_rows():
    return [[o.list_code, o.value_code, o.value_label, o.group_name, o.owner_area_code, "Sí" if o.user_editable else "No", o.sort_order, "Sí" if o.is_active else "No", o.created_by] for o in OptionItem.query.all()]

def _settings_rows():
    return [[s.setting_key, s.setting_value, s.updated_at] for s in AppSetting.query.all()]

def export_repository_workbook():
    template = Path(current_app.config["WORKBOOK_TEMPLATE"])
    output = Path(current_app.config["WORKBOOK_OUTPUT"])
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(template, output)
    wb = load_workbook(output)
    datasets = {
        "ENVIOS": _submission_rows(),
        "RESPUESTAS": _response_rows(),
        "ADJUNTOS": _attachment_rows(),
        "NARRATIVAS": _narrative_rows(),
        "CAT_LISTAS": _list_rows(),
        "CAT_SETTINGS": _settings_rows(),
    }
    for sheet_name, rows in datasets.items():
        if sheet_name in wb.sheetnames:
            _clear_body(wb[sheet_name])
            _write_rows(wb[sheet_name], DATA_SHEET_MAP[sheet_name], rows)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output)
    upload_to_onedrive(output)
    return str(output)

def _graph_token():
    cfg = current_app.config
    if not all([cfg.get("GRAPH_TENANT_ID"), cfg.get("GRAPH_CLIENT_ID"), cfg.get("GRAPH_CLIENT_SECRET")]):
        return None
    token_url = f"https://login.microsoftonline.com/{cfg['GRAPH_TENANT_ID']}/oauth2/v2.0/token"
    data = {
        "client_id": cfg["GRAPH_CLIENT_ID"],
        "client_secret": cfg["GRAPH_CLIENT_SECRET"],
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials"
    }
    resp = requests.post(token_url, data=data, timeout=30)
    resp.raise_for_status()
    return resp.json()["access_token"]

def upload_to_onedrive(filepath):
    cfg = current_app.config
    if not all([cfg.get("GRAPH_DRIVE_ID"), cfg.get("GRAPH_UPLOAD_PATH")]):
        return False, "Graph no configurado"
    token = _graph_token()
    if not token:
        return False, "No se pudo obtener token"
    upload_url = f"https://graph.microsoft.com/v1.0/drives/{cfg['GRAPH_DRIVE_ID']}/root:{cfg['GRAPH_UPLOAD_PATH']}:/content"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    with open(filepath, "rb") as f:
        resp = requests.put(upload_url, headers=headers, data=f.read(), timeout=120)
    if resp.ok:
        return True, resp.json().get("webUrl", "")
    return False, resp.text
